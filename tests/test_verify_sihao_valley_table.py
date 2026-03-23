from __future__ import annotations

import importlib.util
import tempfile
import unittest
import warnings
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPT_PATH = REPO_ROOT / 'scripts' / 'verify_sihao_valley_table.py'
WORKBOOK_PATH = REPO_ROOT / 'Endfield Blueprints (Asia).xlsx'


def load_verify_module():
    spec = importlib.util.spec_from_file_location('verify_sihao_valley_table', SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


verify_module = load_verify_module()


def load_workbook_without_extension_warning(path: Path):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            'ignore',
            message='Unknown extension is not supported and will be removed',
            category=UserWarning,
        )
        return load_workbook(path, read_only=False, data_only=False)


def copy_workbook_to_temp() -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix='sihao-verify-'))
    temp_path = temp_dir / WORKBOOK_PATH.name
    temp_path.write_bytes(WORKBOOK_PATH.read_bytes())
    return temp_path


def find_output_table(workbook):
    for worksheet in workbook.worksheets:
        if verify_module.TABLE_NAME in worksheet.tables:
            return worksheet, worksheet.tables[verify_module.TABLE_NAME]
    raise AssertionError(f"Table {verify_module.TABLE_NAME!r} not found in workbook copy.")


class VerifySihaoValleyTableTests(unittest.TestCase):
    def test_verify_fails_when_output_row_count_differs_from_source_nonblank_rows(self):
        workbook_copy = copy_workbook_to_temp()
        workbook = load_workbook_without_extension_warning(workbook_copy)
        try:
            output_sheet, table = find_output_table(workbook)
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            table.ref = (
                f'{get_column_letter(min_col)}{min_row}:'
                f'{get_column_letter(max_col)}{max_row - 1}'
            )
            workbook.save(workbook_copy)
        finally:
            workbook.close()

        with self.assertRaisesRegex(ValueError, 'row count'):
            verify_module.verify_workbook(workbook_copy, None)

    def test_verify_fails_when_long_note_text_changes_in_output(self):
        workbook_copy = copy_workbook_to_temp()
        workbook = load_workbook_without_extension_warning(workbook_copy)
        try:
            output_sheet, table = find_output_table(workbook)
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            note_column = min_col + verify_module.SOURCE_HEADERS.index('備註')
            for row_index in range(min_row + 1, max_row + 1):
                cell = output_sheet.cell(row=row_index, column=note_column)
                if cell.value and len(str(cell.value)) > 80:
                    cell.value = 'broken note'
                    break
            else:
                self.fail('Expected a long 備註 value in the managed output table.')

            workbook.save(workbook_copy)
        finally:
            workbook.close()

        with self.assertRaisesRegex(ValueError, '備註'):
            verify_module.verify_workbook(workbook_copy, None)

    def test_verify_fails_when_mixed_language_provider_changes_in_output(self):
        workbook_copy = copy_workbook_to_temp()
        workbook = load_workbook_without_extension_warning(workbook_copy)
        try:
            output_sheet, table = find_output_table(workbook)
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            provider_column = min_col + verify_module.SOURCE_HEADERS.index('提供者')
            for row_index in range(min_row + 1, max_row + 1):
                cell = output_sheet.cell(row=row_index, column=provider_column)
                if cell.value and any(ord(character) > 127 for character in str(cell.value)):
                    cell.value = 'provider-only-ascii'
                    break
            else:
                self.fail('Expected a mixed-language 提供者 value in the managed output table.')

            workbook.save(workbook_copy)
        finally:
            workbook.close()

        with self.assertRaisesRegex(ValueError, '提供者'):
            verify_module.verify_workbook(workbook_copy, None)


if __name__ == '__main__':
    unittest.main()
