from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
from typing import Iterable
import sys
import warnings

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet


SOURCE_SHEET_NAME = '四號谷地_1'
TABLE_NAME = 'SihaoValleyTable'
SOURCE_HEADERS = ['時代', '類別', '藍圖名稱', '藍圖代碼', '提供者', '備註']
OUTPUT_HEADERS = SOURCE_HEADERS + ['搜尋文字']
DEFAULT_WORKBOOK = Path(__file__).resolve().parent.parent / 'Endfield Blueprints (Asia).xlsx'


if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Verify the sihao valley managed table workbook state.'
    )
    parser.add_argument(
        'workbook',
        nargs='?',
        default=str(DEFAULT_WORKBOOK),
        help='Workbook to verify. Defaults to the repo workbook.',
    )
    parser.add_argument(
        'baseline_workbook',
        nargs='?',
        default=None,
        help='Optional baseline workbook used to compare source columns A:F.',
    )
    return parser.parse_args()


def require_file(path_text: str) -> Path:
    path = Path(path_text).expanduser().resolve()
    if not path.is_file():
        raise ValueError(f'Workbook not found: {path}')
    return path


def load_excel_workbook(path: Path, read_only: bool):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            'ignore',
            message='Unknown extension is not supported and will be removed',
            category=UserWarning,
        )
        return load_workbook(path, read_only=read_only, data_only=False)


def get_required_sheet(workbook, sheet_name: str) -> Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Source sheet '{sheet_name}' was not found.")
    return workbook[sheet_name]


def find_table_worksheet(workbook, table_name: str) -> tuple[Worksheet, Table]:
    for worksheet in workbook.worksheets:
        if table_name in worksheet.tables:
            return worksheet, worksheet.tables[table_name]
    raise ValueError(f"Managed output table '{table_name}' was not found in the workbook.")


def sheet_headers(worksheet: Worksheet, count: int) -> list[str | None]:
    return [worksheet.cell(row=1, column=column).value for column in range(1, count + 1)]


def table_headers(table: Table) -> list[str]:
    return [column.name for column in table.tableColumns]


def format_headers(headers: Iterable[object]) -> str:
    return '[' + ', '.join('' if value is None else str(value) for value in headers) + ']'


def assert_headers(actual: list[object], expected: list[str], label: str) -> None:
    if actual != expected:
        raise ValueError(
            f"{label} headers did not match. Expected {format_headers(expected)}; "
            f"actual {format_headers(actual)}."
        )


def source_last_row(worksheet: Worksheet) -> int:
    last_row = 1
    for row in range(worksheet.max_row, 0, -1):
        if any(worksheet.cell(row=row, column=column).value is not None for column in range(1, 7)):
            last_row = row
            break
    return last_row


def source_snapshot(worksheet: Worksheet) -> list[tuple[object, ...]]:
    last_row = source_last_row(worksheet)
    rows: list[tuple[object, ...]] = []
    for row in worksheet.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=6, values_only=True):
        rows.append(tuple(row))
    return rows


def compare_source_with_baseline(workbook_path: Path, baseline_path: Path) -> None:
    workbook = load_excel_workbook(workbook_path, read_only=False)
    baseline = load_excel_workbook(baseline_path, read_only=False)
    try:
        source_sheet = get_required_sheet(workbook, SOURCE_SHEET_NAME)
        baseline_sheet = get_required_sheet(baseline, SOURCE_SHEET_NAME)

        actual = source_snapshot(source_sheet)
        expected = source_snapshot(baseline_sheet)
        if actual != expected:
            mismatch_index = next(
                index for index, pair in enumerate(zip(actual, expected), start=1) if pair[0] != pair[1]
            ) if len(actual) == len(expected) else min(len(actual), len(expected)) + 1
            raise ValueError(
                f"Source columns A:F differ from baseline workbook at row {mismatch_index}. "
                f"Workbook: {workbook_path}; baseline: {baseline_path}."
            )
    finally:
        workbook.close()
        baseline.close()


def is_nonblank(value: object) -> bool:
    if value is None:
        return False
    return str(value).strip() != ''


def normalize_text(value: object) -> str:
    if value is None:
        return ''
    return str(value).replace('\r\n', '\n').strip()


def has_ascii_and_non_ascii(value: str) -> bool:
    has_ascii = any(ord(character) < 128 and not character.isspace() for character in value)
    has_non_ascii = any(ord(character) >= 128 for character in value)
    return has_ascii and has_non_ascii


def source_data_rows(worksheet: Worksheet) -> list[tuple[object, ...]]:
    last_row = source_last_row(worksheet)
    rows: list[tuple[object, ...]] = []
    for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=6, values_only=True):
        row_values = tuple(row)
        if any(is_nonblank(value) for value in row_values):
            rows.append(row_values)
    return rows


def output_data_rows(worksheet: Worksheet, table: Table) -> list[tuple[object, ...]]:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows: list[tuple[object, ...]] = []
    for row in worksheet.iter_rows(
        min_row=min_row + 1,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        source_portion = tuple(row[: len(SOURCE_HEADERS)])
        if any(is_nonblank(value) for value in source_portion):
            rows.append(tuple(row))
    return rows


def assert_row_counts_match(source_rows: list[tuple[object, ...]], output_rows: list[tuple[object, ...]]) -> None:
    source_count = len(source_rows)
    output_count = len(output_rows)
    if source_count != output_count:
        raise ValueError(
            f'Source nonblank row count ({source_count}) did not match managed output row count ({output_count}).'
        )


def assert_long_note_text_survives(source_rows: list[tuple[object, ...]], output_rows: list[tuple[object, ...]]) -> None:
    source_notes = Counter(
        normalize_text(row[5])
        for row in source_rows
        if len(normalize_text(row[5])) >= 80
    )
    output_notes = Counter(
        normalize_text(row[5])
        for row in output_rows
        if len(normalize_text(row[5])) >= 80
    )
    missing_notes = [note for note, count in source_notes.items() if output_notes[note] < count]
    if missing_notes:
        raise ValueError(
            'Long 備註 text from the source sheet did not survive in the managed output table.'
        )


def assert_mixed_language_providers_survive(
    source_rows: list[tuple[object, ...]],
    output_rows: list[tuple[object, ...]],
) -> None:
    source_providers = Counter(
        normalize_text(row[4])
        for row in source_rows
        if has_ascii_and_non_ascii(normalize_text(row[4]))
    )
    output_providers = Counter(
        normalize_text(row[4])
        for row in output_rows
        if has_ascii_and_non_ascii(normalize_text(row[4]))
    )
    missing_providers = [provider for provider, count in source_providers.items() if output_providers[provider] < count]
    if missing_providers:
        raise ValueError(
            'Mixed-language 提供者 values from the source sheet did not survive in the managed output table.'
        )


def verify_workbook(workbook_path: Path, baseline_path: Path | None) -> str:
    workbook = load_excel_workbook(workbook_path, read_only=False)
    try:
        source_sheet = get_required_sheet(workbook, SOURCE_SHEET_NAME)
        output_sheet, table = find_table_worksheet(workbook, TABLE_NAME)

        assert_headers(sheet_headers(source_sheet, len(SOURCE_HEADERS)), SOURCE_HEADERS, f"Source sheet '{SOURCE_SHEET_NAME}'")
        assert_headers(table_headers(table), OUTPUT_HEADERS, f"Managed output table '{TABLE_NAME}' on worksheet '{output_sheet.title}'")
        source_rows = source_data_rows(source_sheet)
        output_rows = output_data_rows(output_sheet, table)
        assert_row_counts_match(source_rows, output_rows)
        assert_long_note_text_survives(source_rows, output_rows)
        assert_mixed_language_providers_survive(source_rows, output_rows)
    finally:
        workbook.close()

    if baseline_path is not None:
        compare_source_with_baseline(workbook_path, baseline_path)

    baseline_text = f' and matched baseline {baseline_path}' if baseline_path is not None else ''
    return (
        f"Verified source sheet '{SOURCE_SHEET_NAME}', managed output worksheet '{output_sheet.title}', "
        f"and table '{TABLE_NAME}' in {workbook_path}{baseline_text}."
    )


def main() -> int:
    args = parse_args()

    try:
        workbook_path = require_file(args.workbook)
        baseline_path = require_file(args.baseline_workbook) if args.baseline_workbook else None
        print(verify_workbook(workbook_path, baseline_path))
        return 0
    except Exception as error:
        print(f'ERROR: {error}')
        return 1


if __name__ == '__main__':
    raise SystemExit(main())
